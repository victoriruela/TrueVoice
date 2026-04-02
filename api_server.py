"""
API REST para VibeVoice TTS con FastAPI.
Backend consolidado: generación de audio, voces, narración de carrera, config y proxy Ollama.
"""

import os
import re
import signal
import sys
import uuid
import shutil
import subprocess
import traceback
from dataclasses import asdict
from pathlib import Path
from typing import Optional, List

from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Query, APIRouter
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
import json
import time

import requests as http_client

app = FastAPI(
    title="TrueVoice API",
    description="API REST para generación de audio con VibeVoice y clonación de voz",
    version="1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Rutas del proyecto ──────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).parent.resolve()
VIBEVOICE_REPO = PROJECT_ROOT / "VibeVoice"
# Directorio por defecto: carpeta voices de TrueVoice
DEFAULT_VOICES_DIR = PROJECT_ROOT / "voices"
VIBEVOICE_VOICES_DIR = VIBEVOICE_REPO / "demo" / "voices"
OUTPUTS_DIR = PROJECT_ROOT / "api_outputs"
TEMP_DIR = PROJECT_ROOT / "temp_outputs"
SESSIONS_DIR = PROJECT_ROOT / "race_sessions"
CONFIG_FILE = PROJECT_ROOT / "frontend_config.json"
OUTPUTS_DIR.mkdir(exist_ok=True)
TEMP_DIR.mkdir(exist_ok=True)
SESSIONS_DIR.mkdir(exist_ok=True)
DEFAULT_VOICES_DIR.mkdir(exist_ok=True) # Asegurar que existe

from race_parser import (
    parse_race_file, parse_race_header, generate_ai_intro, generate_ai_descriptions,
    RaceHeader, RaceEvent,
)

# Variable global para mantener compatibilidad con código antiguo que pueda usarla
VOICES_DIR = DEFAULT_VOICES_DIR

# ── Almacenamiento de Progreso ──────────────────────────────────────────────
# Estructura: { "audio_id": { "total": 100, "current": 10, "start_time": float, "last_update": float } }
PROGRESS_STORE = {}
ACTIVE_PROCESSES = {}


def cancel_active_process(progress_id: str) -> bool:
    process = ACTIVE_PROCESSES.get(progress_id)
    if not process:
        return False

    if process.poll() is not None:
        ACTIVE_PROCESSES.pop(progress_id, None)
        return False

    try:
        os.killpg(os.getpgid(process.pid), signal.SIGTERM)
        try:
            process.wait(timeout=5)
        except subprocess.TimeoutExpired:
            os.killpg(os.getpgid(process.pid), signal.SIGKILL)
            process.wait(timeout=2)
        return True
    except Exception as ex:
        print(f"[API] Error cancelando proceso {progress_id}: {ex}")
        return False
    finally:
        ACTIVE_PROCESSES.pop(progress_id, None)
        if progress_id in PROGRESS_STORE:
            PROGRESS_STORE[progress_id]["status"] = "cancelled"
            PROGRESS_STORE[progress_id]["last_update"] = time.time()

print(f"[API] PROJECT_ROOT: {PROJECT_ROOT}")
print(f"[API] DEFAULT_VOICES_DIR: {DEFAULT_VOICES_DIR}")
print(f"[API] OUTPUTS_DIR:  {OUTPUTS_DIR}")


# ── Modelos Pydantic ────────────────────────────────────────────────────────
class GenerateRequest(BaseModel):
    text: str = Field(..., min_length=1, description="Texto a convertir en audio")
    voice_name: str = Field("Alice", description="Nombre de la voz o ruta completa al archivo .wav")
    custom_output_name: Optional[str] = Field(None, description="Nombre personalizado para el archivo de salida")
    output_directory: Optional[str] = Field(None, description="Directorio de salida personalizado")
    audio_id_hint: Optional[str] = Field(None, description="Sugerencia de ID para rastrear progreso")
    model: str = Field("microsoft/VibeVoice-1.5b", description="Modelo HuggingFace")
    output_format: str = Field("wav", description="Formato de salida: wav, mp3, flac, ogg")
    cfg_scale: float = Field(2.0, ge=0.5, le=5.0, description="CFG scale (0.5-5.0)")
    ddpm_steps: int = Field(30, ge=1, le=200, description="Pasos DDPM (1-200)")
    disable_prefill: bool = Field(False, description="Desactivar clonación de voz")


class OutputFileInfo(BaseModel):
    id: str
    filename: str
    path: str
    size: int
    created: float


class VoiceInfo(BaseModel):
    name: str
    filename: str
    alias: Optional[str] = None


class GenerateResponse(BaseModel):
    success: bool
    message: str
    audio_id: Optional[str] = None
    filename: Optional[str] = None
    is_temp: bool = False


class SaveRequest(BaseModel):
    audio_id: str
    output_directory: Optional[str] = None


class ConfigUpdate(BaseModel):
    """Partial config update — only provided keys are merged."""
    selected_voice: Optional[str] = None
    selected_model_name: Optional[str] = None
    selected_model: Optional[str] = None
    output_format: Optional[str] = None
    cfg_scale: Optional[float] = None
    ddpm_steps: Optional[int] = None
    disable_prefill: Optional[bool] = None
    voice_folder_type: Optional[str] = None
    custom_folder_path: Optional[str] = None
    output_folder_type: Optional[str] = None
    custom_output_path: Optional[str] = None
    texts_folder_type: Optional[str] = None
    custom_texts_path: Optional[str] = None
    output_directory: Optional[str] = None
    voice_directory: Optional[str] = None
    ollama_url: Optional[str] = None
    ollama_model: Optional[str] = None
    generation_tasks: Optional[list] = None
    last_text_input: Optional[str] = None
    last_custom_name: Optional[str] = None
    last_race_session: Optional[str] = None
    audio_output_folder: Optional[str] = None
    texts_output_folder: Optional[str] = None


class OllamaGenerateRequest(BaseModel):
    prompt: str = Field(..., min_length=1)
    model: Optional[str] = None
    options: Optional[dict] = None


class RaceSessionSave(BaseModel):
    intro_text: str = ""
    intro_audio: str = ""
    header: dict = {}
    events: list = []
    event_audios: dict = {}


class RaceDescriptionRequest(BaseModel):
    events: list
    ollama_url: Optional[str] = None
    ollama_model: Optional[str] = None


# ── Utilidades ──────────────────────────────────────────────────────────────
DEFAULT_VOICES = {
    "Alice": "en-Alice_woman",
    "Carter": "en-Carter_man",
    "Frank": "en-Frank_man",
    "Mary": "en-Mary_woman_bgm",
    "Maya": "en-Maya_woman",
    "Samuel": "in-Samuel_man",
    "Anchen": "zh-Anchen_man_bgm",
    "Bowen": "zh-Bowen_man",
    "Xinran": "zh-Xinran_woman",
}

ALIAS_REVERSE = {v: k for k, v in DEFAULT_VOICES.items()}


def _resolve_voice(voice_name: str, directory: Optional[str] = None) -> Optional[str]:
    """Resuelve un nombre de voz al nombre exacto del archivo WAV o ruta absoluta."""
    # Si es una ruta absoluta y existe, la usamos directamente
    if os.path.isabs(voice_name) and voice_name.lower().endswith(".wav") and os.path.exists(voice_name):
        return voice_name

    # Determinar qué directorio usar
    target_dir = DEFAULT_VOICES_DIR
    if directory:
        custom_dir = Path(directory)
        if custom_dir.exists() and custom_dir.is_dir():
            target_dir = custom_dir

    if voice_name in DEFAULT_VOICES:
        resolved = DEFAULT_VOICES[voice_name]
        # Primero buscamos en el directorio proporcionado (con el nombre resuelto o el alias)
        if (target_dir / f"{resolved}.wav").exists():
            return str(target_dir / f"{resolved}.wav")
        if (target_dir / f"{voice_name}.wav").exists():
            return str(target_dir / f"{voice_name}.wav")
        # Fallback a VibeVoice original para las voces por defecto
        if (VIBEVOICE_VOICES_DIR / f"{resolved}.wav").exists():
            return str(VIBEVOICE_VOICES_DIR / f"{resolved}.wav")

    if (target_dir / f"{voice_name}.wav").exists():
        return str(target_dir / f"{voice_name}.wav")

    voice_lower = voice_name.lower()
    for wav in target_dir.glob("*.wav"):
        if voice_lower == wav.stem.lower(): # Coincidencia exacta primero
            return str(wav)
    
    for wav in target_dir.glob("*.wav"):
        if voice_lower in wav.stem.lower(): # Coincidencia parcial después
            return str(wav)

    return None


# ── Endpoints ───────────────────────────────────────────────────────────────
@app.get("/")
def root():
    return {"status": "ok", "service": "TrueVoice API", "version": "1.0.0"}


@app.get("/voices", response_model=list[VoiceInfo])
def list_voices(directory: Optional[str] = None):
    """Lista todas las voces disponibles en el directorio especificado o el predeterminado."""
    voices = []
    target_dir = DEFAULT_VOICES_DIR
    
    if directory:
        custom_dir = Path(directory)
        if custom_dir.exists() and custom_dir.is_dir():
            target_dir = custom_dir
        else:
            # Si el directorio no existe, devolvemos lista vacía
            return []

    for wav in sorted(target_dir.glob("*.wav")):
        alias = ALIAS_REVERSE.get(wav.stem)
        voices.append(VoiceInfo(name=wav.stem, filename=wav.name, alias=alias))
    
    return voices


@app.post("/generate", response_model=GenerateResponse)
def generate_audio(req: GenerateRequest, voice_directory: Optional[str] = None):
    """Genera audio a partir de texto usando VibeVoice."""
    try:
        print(f"[API] Solicitud de generación recibida. Voz: {req.voice_name}, Directorio: {voice_directory}")
        # Valida la voz (pasamos el directorio si existe)
        resolved_path = _resolve_voice(req.voice_name, voice_directory)
        if resolved_path is None:
            raise HTTPException(404, f"Voz '{req.voice_name}' no encontrada en el directorio especificado ({voice_directory or 'default'}).")

        # Valida formato
        fmt = req.output_format.lower().lstrip(".")
        if fmt not in ("wav", "mp3", "flac", "ogg"):
            raise HTTPException(400, f"Formato '{fmt}' no soportado. Usa: wav, mp3, flac, ogg")

        # Determinar directorio de salida (ahora siempre es el temporal inicialmente)
        target_output_dir = TEMP_DIR
        
        # Guardar dónde debería ir finalmente para referencia
        final_target_dir = OUTPUTS_DIR
        if req.output_directory:
            custom_out = Path(req.output_directory)
            if custom_out.exists() and custom_out.is_dir():
                final_target_dir = custom_out
            else:
                try:
                    custom_out.mkdir(parents=True, exist_ok=True)
                    final_target_dir = custom_out
                except Exception as e:
                    print(f"[API] No se pudo crear el directorio de salida personalizado: {e}")

        # Genera un ID único para este audio
        if req.custom_output_name:
            # Sanitizar nombre (quitar extensión si la puso)
            base_name = Path(req.custom_output_name).stem
            # Lógica de evitar duplicados en el destino final
            final_name = f"{base_name}.{fmt}"
            counter = 1
            while (final_target_dir / final_name).exists():
                final_name = f"{base_name}_{counter}.{fmt}"
                counter += 1
            output_file = target_output_dir / final_name
            audio_id = Path(final_name).stem # Usamos el nombre base como ID si es personalizado
        else:
            # Comportamiento por defecto (UUID)
            audio_id = uuid.uuid4().hex[:12]
            output_file = target_output_dir / f"{audio_id}.{fmt}"

        # Si el cliente nos da una sugerencia (hint) para el progreso, la usamos o añadimos a la store
        progress_id = req.audio_id_hint if req.audio_id_hint else audio_id

        # Si el ID ya existe (por un reintento o coincidencia), limpiar progreso previo
        if progress_id in PROGRESS_STORE:
            print(f"[API] Limpiando progreso previo para {progress_id}")
            del PROGRESS_STORE[progress_id]

        # Inicializar progreso
        PROGRESS_STORE[progress_id] = {
            "total": 0,
            "current": 0,
            "start_time": time.time(),
            "last_update": time.time(),
            "status": "starting"
        }

        # Construye el comando
        vibevoice_script = PROJECT_ROOT / "vibevoice_app.py"
        # MODO TEST SI SE ACTIVA: 
        # vibevoice_script = PROJECT_ROOT / "test_gen_sim.py"
        cmd = [
            sys.executable, str(vibevoice_script),
            "--text", req.text,
            "--voice-name", resolved_path, # Pasamos la ruta resuelta
            "--model", req.model,
            "--output", str(output_file),
            "--cfg-scale", str(req.cfg_scale),
            "--ddpm-steps", str(req.ddpm_steps),
        ]
        if req.disable_prefill:
            cmd.append("--disable-prefill")

        print(f"\n{'='*60}")
        print(f"[API] Generando audio...")
        print(f"[API] Voz: {req.voice_name} → {resolved_path}")
        print(f"[API] Output: {output_file}")
        print(f"[API] Comando: {' '.join(cmd[:6])}...")
        print(f"{'='*60}")

        # Ejecución con captura de progreso
        process = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            # encoding="utf-8",
            errors="replace",
            cwd=str(PROJECT_ROOT),
            bufsize=1, # Line buffered
            start_new_session=True,
        )
        ACTIVE_PROCESSES[progress_id] = process

        stdout_acc = []
        stderr_acc = []

        # Usar un hilo para leer stderr y evitar bloqueos de buffer
        import threading
        def enqueue_stderr(pipe, acc, prog_id):
            try:
                for line in iter(pipe.readline, ''):
                    acc.append(line)
                    clean_line = line.strip()
                    # También buscamos progreso en stderr por si acaso
                    if clean_line.startswith("PROGRESS_STEP:"):
                        try:
                            current = int(clean_line.split(":")[1])
                            PROGRESS_STORE[prog_id]["current"] = current
                            PROGRESS_STORE[prog_id]["last_update"] = time.time()
                        except: pass
                    # print(f"[API-ERR] {line.strip()}")
                    print(f"[API-ERR] {line.strip()}", flush=True)
            finally:
                pipe.close()

        stderr_thread = threading.Thread(target=enqueue_stderr, args=(process.stderr, stderr_acc, progress_id), daemon=True)
        stderr_thread.start()

        # Leer stdout línea a línea para el progreso
        while True:
            line = process.stdout.readline()
            if not line and process.poll() is not None:
                break
            if line:
                stdout_acc.append(line)
                clean_line = line.strip()
                # print(f"[API-OUT] {clean_line}") # Debug: Ver todo lo que sale
                if clean_line.startswith("PROGRESS_START:"):
                    try:
                        total = int(clean_line.split(":")[1])
                        PROGRESS_STORE[progress_id]["total"] = total
                        PROGRESS_STORE[progress_id]["status"] = "generating"
                        print(f"[API] Progreso detectado: Total {total}")
                    except: pass
                elif clean_line.startswith("PROGRESS_STEP:"):
                    try:
                        current = int(clean_line.split(":")[1])
                        PROGRESS_STORE[progress_id]["current"] = current
                        PROGRESS_STORE[progress_id]["last_update"] = time.time()
                        # print(f"[API] Progreso detectado: Step {current}/{PROGRESS_STORE[progress_id]['total']}")
                    except: pass
                # print(f"[API-PROC] {clean_line}")

        # Capturar stderr restante (aunque ya lo hace el hilo, esperamos a que termine)
        stderr_thread.join(timeout=5)
        ACTIVE_PROCESSES.pop(progress_id, None)
        
        # Log completo para debug
        print(f"[API] Return code: {process.returncode}")
        if process.returncode != 0:
            print("[API] === STDOUT ===")
            print("".join(stdout_acc[-50:]), flush=True)
        
        if process.returncode != 0:
            PROGRESS_STORE[progress_id]["status"] = "failed"
            detail = "".join(stderr_acc)[-500:] or "".join(stdout_acc)[-500:] or "Sin salida"
            raise HTTPException(500, f"Error en vibevoice_app.py (code {process.returncode}):\n{detail}")

        if not output_file.exists():
            PROGRESS_STORE[progress_id]["status"] = "failed"
            raise HTTPException(500, f"El script terminó OK pero no generó el archivo: {output_file}")

        PROGRESS_STORE[progress_id]["status"] = "completed"
        PROGRESS_STORE[progress_id]["current"] = PROGRESS_STORE[progress_id]["total"]
        print(f"[API] ✅ Audio generado: {output_file} ({output_file.stat().st_size} bytes)")

        return GenerateResponse(
            success=True,
            message="Audio generado correctamente (Pendiente de guardar)",
            audio_id=audio_id,
            filename=output_file.name,
            is_temp=True
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"[API] ❌ Excepción no controlada: {e}")
        traceback.print_exc()
        raise HTTPException(500, f"Error interno: {str(e)}")


@app.get("/audio/{audio_id}")
def download_audio(audio_id: str, directory: Optional[str] = None):
    """Descarga un audio generado por su ID."""
    # Primero buscar en temporal
    matches = list(TEMP_DIR.glob(f"{audio_id}.*"))
    if matches:
        return FileResponse(matches[0], media_type="audio/mpeg", filename=matches[0].name)

    target_dir = OUTPUTS_DIR
    if directory:
        custom_dir = Path(directory)
        if custom_dir.exists() and custom_dir.is_dir():
            target_dir = custom_dir

    matches = list(target_dir.glob(f"{audio_id}.*"))
    if not matches:
        raise HTTPException(404, f"Audio '{audio_id}' no encontrado")
    return FileResponse(matches[0], media_type="audio/mpeg", filename=matches[0].name)


@app.post("/confirm_save")
def confirm_save(req: SaveRequest):
    """Mueve un audio de la carpeta temporal a la definitiva."""
    matches = list(TEMP_DIR.glob(f"{req.audio_id}.*"))
    if not matches:
        raise HTTPException(404, f"Audio temporal '{req.audio_id}' no encontrado")
    
    temp_file = matches[0]
    target_dir = OUTPUTS_DIR
    if req.output_directory:
        custom_dir = Path(req.output_directory)
        if custom_dir.exists() and custom_dir.is_dir():
            target_dir = custom_dir
        else:
            custom_dir.mkdir(parents=True, exist_ok=True)
            target_dir = custom_dir

    final_path = target_dir / temp_file.name
    
    # Manejar colisiones por si acaso
    if final_path.exists():
        base = temp_file.stem
        ext = temp_file.suffix
        counter = 1
        while (target_dir / f"{base}_{counter}{ext}").exists():
            counter += 1
        final_path = target_dir / f"{base}_{counter}{ext}"

    try:
        shutil.move(str(temp_file), str(final_path))
        print(f"[API] Audio guardado permanentemente: {final_path}")
        return {"success": True, "message": f"Audio guardado en {final_path}", "filename": final_path.name}
    except Exception as e:
        raise HTTPException(500, f"Error al guardar el archivo: {str(e)}")


@app.post("/cleanup_temp")
def cleanup_temp():
    """Cancela procesos activos y borra todos los archivos de la carpeta temporal."""
    cancelled = 0
    for progress_id in list(ACTIVE_PROCESSES.keys()):
        if cancel_active_process(progress_id):
            cancelled += 1

    count = 0
    for f in TEMP_DIR.glob("*.*"):
        try:
            f.unlink()
            count += 1
        except: pass
    print(f"[API] Limpieza temporal: {cancelled} procesos cancelados, {count} archivos borrados")
    return {"success": True, "count": count, "cancelled": cancelled}


@app.get("/progress/{audio_id}")
def get_progress(audio_id: str):
    """Obtiene el progreso de una generación específica."""
    if audio_id not in PROGRESS_STORE:
        raise HTTPException(404, "ID de audio no encontrado")
    return PROGRESS_STORE[audio_id]


@app.get("/outputs", response_model=list[OutputFileInfo])
def list_outputs(directory: Optional[str] = None):
    """Lista todos los audios generados en el directorio especificado."""
    target_dir = OUTPUTS_DIR
    if directory:
        custom_dir = Path(directory)
        if custom_dir.exists() and custom_dir.is_dir():
            target_dir = custom_dir

    files = []
    # Extensiones de audio comunes
    for ext in ("*.wav", "*.mp3", "*.flac", "*.ogg"):
        for f in target_dir.glob(ext):
            files.append(OutputFileInfo(
                id=f.stem,
                filename=f.name,
                path=str(f),
                size=f.stat().st_size,
                created=f.stat().st_mtime
            ))
    
    # Ordenar por fecha de creación descendente
    files.sort(key=lambda x: x.created, reverse=True)
    return files


@app.delete("/outputs/delete")
def delete_outputs(filenames: list[str] = Query(...), directory: Optional[str] = None):
    """Borra uno o varios archivos de audio."""
    target_dir = OUTPUTS_DIR
    if directory:
        custom_dir = Path(directory)
        if custom_dir.exists() and custom_dir.is_dir():
            target_dir = custom_dir

    deleted = []
    errors = []
    
    for filename in filenames:
        file_path = target_dir / filename
        if file_path.exists() and file_path.is_file():
            try:
                os.remove(file_path)
                deleted.append(filename)
            except Exception as e:
                errors.append(f"Error al borrar {filename}: {str(e)}")
        else:
            errors.append(f"Archivo no encontrado: {filename}")
            
    return {"deleted": deleted, "errors": errors}


@app.post("/voices/upload", response_model=VoiceInfo)
async def upload_voice(
    voice_name: str = Form(..., description="Nombre para la nueva voz"),
    audio_file: UploadFile = File(..., description="Archivo de audio WAV"),
):
    """Sube un archivo de audio como nueva voz personalizada."""
    if not voice_name.strip():
        raise HTTPException(400, "El nombre de la voz no puede estar vacío")

    temp_path = OUTPUTS_DIR / f"upload_{uuid.uuid4().hex[:8]}.wav"
    content = await audio_file.read()
    temp_path.write_bytes(content)

    cmd = [
        sys.executable, str(PROJECT_ROOT / "vibevoice_app.py"),
        "--clone-voice", str(temp_path),
        "--voice-name", voice_name,
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=60, cwd=str(PROJECT_ROOT))

    if temp_path.exists():
        temp_path.unlink()

    voice_path = DEFAULT_VOICES_DIR / f"{voice_name}.wav"
    if not voice_path.exists():
        raise HTTPException(500, f"Error procesando voz: {result.stderr[-300:]}")

    return VoiceInfo(name=voice_name, filename=voice_path.name)


@app.delete("/voices/{voice_name}")
def delete_voice(voice_name: str):
    """Elimina una voz personalizada (solo del directorio por defecto)."""
    voice_path = DEFAULT_VOICES_DIR / f"{voice_name}.wav"
    if not voice_path.exists():
        raise HTTPException(404, f"Voz '{voice_name}' no encontrada")

    if voice_name in ALIAS_REVERSE:
        raise HTTPException(403, f"No se puede eliminar la voz predeterminada '{voice_name}'")

    voice_path.unlink()
    return {"success": True, "message": f"Voz '{voice_name}' eliminada"}


@app.get("/models")
def list_models():
    """Lista los modelos disponibles."""
    return [
        {"id": "microsoft/VibeVoice-1.5b", "name": "VibeVoice 1.5B (recomendado)", "size": "~6GB"},
        {"id": "microsoft/VibeVoice-7b", "name": "VibeVoice 7B (alta calidad)", "size": "~28GB"},
    ]


# ══════════════════════════════════════════════════════════════════════════════
# CONFIG ENDPOINTS
# ══════════════════════════════════════════════════════════════════════════════

def _load_config() -> dict:
    """Read frontend_config.json, return empty dict on error."""
    if CONFIG_FILE.exists():
        try:
            return json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def _save_config(config: dict) -> None:
    """Write config dict to frontend_config.json."""
    CONFIG_FILE.write_text(json.dumps(config, indent=4, ensure_ascii=False), encoding="utf-8")


@app.get("/config")
def get_config():
    """Return the full frontend configuration."""
    return _load_config()


@app.put("/config")
def update_config(update: ConfigUpdate):
    """Partial-merge update of frontend configuration. Only provided (non-null) keys are written."""
    config = _load_config()
    patch = {k: v for k, v in update.model_dump().items() if v is not None}
    config.update(patch)
    _save_config(config)
    return config


# ══════════════════════════════════════════════════════════════════════════════
# OLLAMA PROXY ENDPOINTS
# ══════════════════════════════════════════════════════════════════════════════

def _get_ollama_url() -> str:
    config = _load_config()
    return config.get("ollama_url", "http://localhost:11434")


@app.get("/ollama/models")
def ollama_list_models():
    """Proxy: list available Ollama models."""
    ollama_url = _get_ollama_url()
    try:
        resp = http_client.get(f"{ollama_url}/api/tags", timeout=15)
        resp.raise_for_status()
        models = resp.json().get("models", [])
        return [m.get("name", m.get("model", "unknown")) for m in models]
    except http_client.ConnectionError:
        raise HTTPException(502, f"No se pudo conectar a Ollama en {ollama_url}")
    except Exception as e:
        raise HTTPException(502, f"Error al consultar Ollama: {e}")


@app.post("/ollama/generate")
def ollama_generate(req: OllamaGenerateRequest):
    """Proxy: generate text via Ollama."""
    config = _load_config()
    ollama_url = config.get("ollama_url", "http://localhost:11434")
    model = req.model or config.get("ollama_model", "llama3.2")

    payload = {"model": model, "prompt": req.prompt, "stream": False}
    if req.options:
        payload["options"] = req.options

    try:
        resp = http_client.post(f"{ollama_url}/api/generate", json=payload, timeout=120)
        resp.raise_for_status()
        data = resp.json()
        return {"text": data.get("response", "").strip(), "model": model}
    except http_client.ConnectionError:
        raise HTTPException(502, f"No se pudo conectar a Ollama en {ollama_url}")
    except Exception as e:
        raise HTTPException(502, f"Error en Ollama: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# DIRECTORY BROWSING ENDPOINTS
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/browse/drives")
def browse_drives():
    """List mounted drives (useful inside Docker on Windows host)."""
    drives = []
    mnt = Path("/mnt")
    if mnt.exists():
        for child in sorted(mnt.iterdir()):
            if child.is_dir() and len(child.name) == 1 and child.name.isalpha():
                drives.append(child.name)
    if not drives:
        drives = ["c"]
    return drives


@app.get("/browse/folders")
def browse_folders(path: str = Query(..., description="Absolute path to list")):
    """List subdirectories at the given path."""
    target = Path(path)
    if not target.exists() or not target.is_dir():
        raise HTTPException(404, f"Directorio no encontrado: {path}")

    items = []
    try:
        for child in sorted(target.iterdir()):
            if child.is_dir():
                items.append({"name": child.name, "path": str(child)})
    except PermissionError:
        raise HTTPException(403, f"Sin permisos para leer: {path}")

    return {"current": str(target), "parent": str(target.parent), "folders": items}


# ══════════════════════════════════════════════════════════════════════════════
# RACE NARRATION ENDPOINTS
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/race/parse")
async def race_parse(file: UploadFile = File(...)):
    """Parse an rFactor2 XML results file. Returns header + events."""
    content = (await file.read()).decode("utf-8", errors="replace")
    header = parse_race_header(content)
    events = parse_race_file(content)
    return {
        "header": asdict(header),
        "events": [asdict(e) for e in events],
    }


@app.post("/race/intro")
def race_generate_intro(header: dict, ollama_url: Optional[str] = None, ollama_model: Optional[str] = None):
    """Generate a race introduction text via Ollama."""
    config = _load_config()
    url = ollama_url or config.get("ollama_url", "http://localhost:11434")
    model = ollama_model or config.get("ollama_model", "llama3.2")

    rh = RaceHeader(
        track_event=header.get("track_event", ""),
        track_length=header.get("track_length", 0),
        race_laps=header.get("race_laps", 0),
        num_drivers=header.get("num_drivers", 0),
        grid_order=header.get("grid_order", []),
        intro_text=header.get("intro_text", ""),
    )
    result = generate_ai_intro(rh, url, model)
    return {"intro_text": result.intro_text}


@app.post("/race/descriptions")
def race_generate_descriptions(req: RaceDescriptionRequest):
    """Generate AI descriptions for race events via Ollama."""
    config = _load_config()
    url = req.ollama_url or config.get("ollama_url", "http://localhost:11434")
    model = req.ollama_model or config.get("ollama_model", "llama3.2")

    events = []
    for e in req.events:
        events.append(RaceEvent(
            lap=e.get("lap", 0),
            timestamp=e.get("timestamp", 0.0),
            event_type=e.get("event_type", 1),
            summary=e.get("summary", ""),
            description=e.get("description", ""),
        ))

    result_events = generate_ai_descriptions(events, url, model)
    return {"events": [asdict(e) for e in result_events]}


@app.get("/race/sessions")
def race_list_sessions():
    """List saved race narration sessions."""
    sessions = []
    for f in sorted(SESSIONS_DIR.glob("*.json")):
        sessions.append({"name": f.stem, "filename": f.name, "size": f.stat().st_size, "modified": f.stat().st_mtime})
    return sessions


@app.get("/race/sessions/{name}")
def race_get_session(name: str):
    """Load a saved race session by name."""
    path = SESSIONS_DIR / f"{name}.json"
    if not path.exists():
        raise HTTPException(404, f"Sesión '{name}' no encontrada")
    return json.loads(path.read_text(encoding="utf-8"))


@app.post("/race/sessions/{name}")
def race_save_session(name: str, session: RaceSessionSave):
    """Save a race narration session."""
    path = SESSIONS_DIR / f"{name}.json"
    path.write_text(json.dumps(session.model_dump(), indent=2, ensure_ascii=False), encoding="utf-8")
    return {"success": True, "name": name}


@app.delete("/race/sessions/{name}")
def race_delete_session(name: str):
    """Delete a saved race session."""
    path = SESSIONS_DIR / f"{name}.json"
    if not path.exists():
        raise HTTPException(404, f"Sesión '{name}' no encontrada")
    path.unlink()
    return {"success": True}


@app.get("/race/sessions/{name}/excel")
def race_export_excel(name: str):
    """Export a race session to Excel and return the file."""
    session_path = SESSIONS_DIR / f"{name}.json"
    if not session_path.exists():
        raise HTTPException(404, f"Sesión '{name}' no encontrada")

    session = json.loads(session_path.read_text(encoding="utf-8"))
    events = session.get("events", [])

    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(500, "openpyxl no está instalado en el servidor")

    wb = Workbook()
    ws = wb.active
    ws.title = "Eventos"
    ws.append(["Vuelta", "Timestamp", "Tipo", "Resumen", "Descripción IA", "Audio"])

    type_labels = {1: "Adelantamiento", 2: "Choque entre pilotos", 3: "Choque contra muro", 4: "Penalización", 5: "Entrada a boxes"}
    event_audios = session.get("event_audios", {})

    for i, ev in enumerate(events):
        lap = ev.get("lap", "")
        ts = ev.get("timestamp", 0)
        minutes, secs = divmod(int(ts), 60)
        hours, minutes = divmod(minutes, 60)
        ts_str = f"{hours:02d}:{minutes:02d}:{secs:02d}"
        etype = type_labels.get(ev.get("event_type", 0), f"Tipo {ev.get('event_type', '?')}")
        ws.append([lap, ts_str, etype, ev.get("summary", ""), ev.get("description", ""), event_audios.get(str(i), "")])

    # Write intro as first row if present
    intro = session.get("intro_text", "")
    if intro:
        ws.insert_rows(2)
        ws.cell(row=2, column=1, value=0)
        ws.cell(row=2, column=2, value="00:00:00")
        ws.cell(row=2, column=3, value="Introducción")
        ws.cell(row=2, column=5, value=intro)
        ws.cell(row=2, column=6, value=session.get("intro_audio", ""))

    import io
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_name = re.sub(r'[^\w\s\-]', '', name).replace(' ', '_')
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'},
    )
